#!/usr/bin/env python3
"""處理紐西蘭 Medsafe/PHARMAC 藥品資料

從 PHARMAC 下載社區藥品時程表 (Community Pharmaceutical Schedule) Excel，
或處理手動下載的 NZULM CSV，並轉換為 JSON 格式。

使用方法:
    uv run python scripts/process_fda_data.py

資料來源:
    主要: PHARMAC Community Pharmaceutical Schedule (Excel 自動下載)
          https://schedule.pharmac.govt.nz/latest/CPSReporting.xlsx
    備用: NZULM (需透過 email 申請)
          https://info.nzulm.org.nz/data-access

產生檔案:
    data/raw/nz_fda_drugs.json
"""

import json
from pathlib import Path

import pandas as pd
import requests
import yaml


def load_config() -> dict:
    """載入欄位映射設定"""
    config_path = Path(__file__).parent.parent / "config" / "fields.yaml"
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def download_pharmac_schedule(output_path: Path) -> Path:
    """從 PHARMAC 下載社區藥品時程表 Excel

    PHARMAC 提供每月更新的 CPSReporting.xlsx，包含紐西蘭已資助的藥品清單。

    Args:
        output_path: Excel 輸出路徑

    Returns:
        下載的檔案路徑
    """
    config = load_config()
    ds = config["data_source"]
    url = ds["url"]

    print("正在從 PHARMAC 下載 Community Pharmaceutical Schedule...")
    print(f"下載 URL: {url}")
    print()

    try:
        response = requests.get(url, timeout=120, headers={
            "User-Agent": "Mozilla/5.0 (compatible; TxGNN/1.0; research)",
        })
        response.raise_for_status()

        content_type = response.headers.get("Content-Type", "")
        if "html" in content_type.lower():
            raise ValueError("返回 HTML 頁面而非 Excel 檔案")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(response.content)

        size_mb = output_path.stat().st_size / 1024 / 1024
        print(f"下載完成: {output_path}")
        print(f"檔案大小: {size_mb:.1f} MB")

        if size_mb < 0.01:
            output_path.unlink(missing_ok=True)
            raise ValueError(f"檔案過小 ({size_mb:.3f} MB)，可能不是有效資料")

        return output_path

    except (requests.RequestException, ValueError) as e:
        print(f"\nPHARMAC 下載失敗: {e}")
        print()

        # 嘗試存檔版本
        archive_url = ds.get("archive_url", "")
        if archive_url:
            print(f"嘗試存檔版本: {archive_url}")
            try:
                response = requests.get(archive_url, timeout=120, headers={
                    "User-Agent": "Mozilla/5.0 (compatible; TxGNN/1.0; research)",
                })
                response.raise_for_status()

                output_path.parent.mkdir(parents=True, exist_ok=True)
                with open(output_path, "wb") as f:
                    f.write(response.content)

                print(f"存檔版本下載完成: {output_path}")
                return output_path
            except requests.RequestException as e2:
                print(f"存檔版本也失敗: {e2}")

        raise FileNotFoundError(
            f"無法自動下載 PHARMAC 藥品時程表\n\n"
            f"請手動取得資料：\n\n"
            f"方法 A: PHARMAC Community Pharmaceutical Schedule (推薦)\n"
            f"  1. 前往 https://schedule.pharmac.govt.nz/\n"
            f"  2. 下載 CPSReporting.xlsx\n"
            f"  3. 將檔案放置於: {output_path}\n\n"
            f"方法 B: NZULM 月度 CSV\n"
            f"  1. 前往 https://info.nzulm.org.nz/data-access\n"
            f"  2. 申請並下載月度 CSV\n"
            f"  3. 將 CSV 檔案放置於 {output_path.parent}/\n\n"
            f"支援格式: .xlsx, .csv"
        )


def find_existing_data(raw_dir: Path) -> Path | None:
    """在 raw 目錄中尋找已存在的資料檔案

    Args:
        raw_dir: data/raw/ 目錄路徑

    Returns:
        找到的檔案路徑，或 None
    """
    # 嘗試常見檔名
    candidates = [
        raw_dir / "CPSReporting.xlsx",
        raw_dir / "pharmac_schedule.xlsx",
        raw_dir / "nzulm.csv",
        raw_dir / "NZULM.csv",
        raw_dir / "medsafe.csv",
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    # 搜尋任何 Excel 或 CSV
    for pattern in ["*.xlsx", "*.xls", "*.csv"]:
        files = list(raw_dir.glob(pattern))
        if files:
            return files[0]

    return None


def process_pharmac_excel(excel_path: Path, output_path: Path) -> Path:
    """處理 PHARMAC Excel 並轉換為 JSON

    Args:
        excel_path: Excel 檔案路徑
        output_path: JSON 輸出路徑

    Returns:
        輸出檔案路徑
    """
    config = load_config()

    print(f"讀取資料檔案: {excel_path}")

    suffix = excel_path.suffix.lower()

    if suffix == ".csv":
        try:
            df = pd.read_csv(excel_path, encoding="utf-8", dtype=str, on_bad_lines="skip")
        except UnicodeDecodeError:
            df = pd.read_csv(excel_path, encoding="latin-1", dtype=str, on_bad_lines="skip")
    elif suffix in (".xlsx", ".xls"):
        # PHARMAC CPSReporting.xlsx has metadata rows at the top.
        # Auto-detect: find the correct sheet and header row by scanning
        # for known column names (Chemical, Pharmacode, Brand, etc.)
        import openpyxl
        target_headers = {"Chemical", "Pharmacode", "Brand", "Presentation", "Subsidy", "Price"}

        sheet_name = 0
        header_row = 0  # 0-indexed for pandas

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            # Prefer "Community Medicines" sheet, then scan all sheets
            preferred_sheets = ["Community Medicines", "Combined Schedule information"]
            ordered_sheets = []
            for ps in preferred_sheets:
                if ps in wb.sheetnames:
                    ordered_sheets.append(ps)
            for sn in wb.sheetnames:
                if sn not in ordered_sheets:
                    ordered_sheets.append(sn)

            found = False
            for sn in ordered_sheets:
                ws = wb[sn]
                for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
                    cell_values = {str(c).strip() for c in row if c is not None}
                    if len(target_headers & cell_values) >= 3:
                        sheet_name = sn
                        header_row = row_idx
                        print(f"找到資料標題列: sheet='{sn}', row={row_idx} (0-indexed)")
                        found = True
                        break
                if found:
                    break
            wb.close()
        except Exception as e:
            print(f"自動偵測標題列失敗 ({e})，使用預設值")

        try:
            df = pd.read_excel(
                excel_path,
                engine="openpyxl",
                dtype=str,
                sheet_name=sheet_name,
                header=header_row,
            )
        except Exception:
            df = pd.read_excel(
                excel_path,
                dtype=str,
                sheet_name=sheet_name,
                header=header_row,
            )
    else:
        raise ValueError(f"不支援的檔案格式: {suffix}")

    print(f"原始資料筆數: {len(df)}")
    print(f"欄位: {', '.join(df.columns.tolist())}")

    # 清理資料
    df = df.fillna("")

    # 轉換為 JSON
    data = df.to_dict(orient="records")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"儲存至: {output_path}")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"完成！共 {len(data)} 筆藥品資料")

    # 顯示統計
    print_statistics(df, config)

    return output_path


def print_statistics(df: pd.DataFrame, config: dict):
    """印出資料統計"""
    fm = config["field_mapping"]
    status_field = fm["status"]
    ingredients_field = fm["ingredients"]

    print("\n" + "=" * 50)
    print("資料統計")
    print("=" * 50)

    if status_field in df.columns:
        print(f"\n註冊狀態分布:")
        status_counts = df[status_field].value_counts()
        for status, count in status_counts.items():
            print(f"  {status}: {count:,}")
    else:
        print(f"\n狀態欄位 '{status_field}' 不存在，可用欄位:")
        for col in df.columns[:10]:
            print(f"  - {col}")

    if ingredients_field in df.columns:
        with_ingredients = (df[ingredients_field] != "").sum()
        if len(df) > 0:
            print(f"\n有活性成分: {with_ingredients:,} ({with_ingredients/len(df)*100:.1f}%)")


def main():
    print("=" * 60)
    print("處理紐西蘭 Medsafe/PHARMAC 藥品資料")
    print("=" * 60)
    print()

    base_dir = Path(__file__).parent.parent
    raw_dir = base_dir / "data" / "raw"
    pharmac_path = raw_dir / "CPSReporting.xlsx"
    output_path = raw_dir / "nz_fda_drugs.json"

    # 確保 raw 目錄存在
    raw_dir.mkdir(parents=True, exist_ok=True)

    # 尋找已存在的檔案
    existing = find_existing_data(raw_dir)
    if existing:
        print(f"使用已存在的資料檔案: {existing}")
        data_path = existing
    else:
        # 自動下載 PHARMAC Schedule
        data_path = download_pharmac_schedule(pharmac_path)

    # 處理資料
    process_pharmac_excel(data_path, output_path)

    print()
    print("下一步: 準備詞彙表資料")
    print("  uv run python scripts/prepare_external_data.py")


if __name__ == "__main__":
    main()
